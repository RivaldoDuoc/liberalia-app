# views.py
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from datetime import date, datetime
from decimal import Decimal
from openpyxl.utils import get_column_letter


# Excluir campos internos del export (p. ej. id, nombre de fichero de imagen)
EXCLUDE_KEYS = {"id", "codigo_imagen"}


def _format_value(v):
    if v is None:
        return "-"
    if isinstance(v, (date, datetime)):
        return v.strftime("%d-%m-%Y")
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, (list, tuple)):
        s = ", ".join(map(str, v))
        return s if s else "-"
    if isinstance(v, dict):
        # join simple dict values for readability
        s = ", ".join(str(x) for x in v.values())
        return s if s else "-"
    # Convert empty strings to '-'
    if isinstance(v, str) and v.strip() == "":
        return "-"
    return v


def exportar_excel(data):
    # data: lista de dicts (por ejemplo resultado de queryset.values())
    # Si no hay datos, devolvemos un excel vacío con encabezado genérico
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado de la búsqueda"

    # Determinar keys/headers dinámicamente a partir del primer registro
    if data and isinstance(data, (list, tuple)) and len(data) > 0 and isinstance(data[0], dict):
        keys = [k for k in list(data[0].keys()) if k not in EXCLUDE_KEYS]
        # Prettify header names: replace __ or . with space and title-case
        headers = [k.replace("__", " ").replace('.', ' ').replace('_', ' ').title() for k in keys]
    else:
        # si no hay dicts, exportamos en una sola columna textual
        keys = None
        headers = ["Resultado"]

    # Nombres de columnas
    ws.append(headers)

    # Estilos
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="7b1e2d", end_color="7b1e2d", fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Filas
    if keys is None:
        for item in data:
            ws.append([_format_value(item)])
    else:
        for item in data:
            row = []
            for k in keys:
                # item is a dict; get the value for this key
                v = item.get(k)
                row.append(_format_value(v))
            ws.append(row)
    # Aplicar estilos: bordes a todas las celdas con información y estilo a encabezados
    max_row = ws.max_row
    max_col = ws.max_column

    # Estilizar encabezados (fila 1)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Aplicar borde negro a celdas que contienen datos (desde fila 2)
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            # solo aplicar borde si hay valor (no vacío)
            if cell.value not in (None, ""):
                cell.border = thin_border
    # Ajustar ancho de columnas según el contenido (incluye encabezado)
    for col in range(1, max_col + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            # contar longitud del texto tal como aparecerá
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
        # corrección para que no quede muy ajustado y límite máximo razonable
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="resultado_busqueda.xlsx"'
    wb.save(response)
    return response
